#!/usr/bin/env python3
"""EON fiyat motoru (v4 grid) — saf yeniden-uretim.

Fiyat panelin DISINDA belirlenir (0119 `externalPricing`, 0120 ayna). Bu modul
motoru YENIDEN URETMEZ, motorun KENDI formulunu birebir cozer ve dogrular:
v4 xlsx'in 858 hucresinin hepsi sifir sapmayla yeniden hesaplanir (asagidaki
`verify_against_grid`). Amac tam-beden ARALARINI (yarim bedenler) ayni formulle
doldurabilmek — grid yalnizca 13 tam beden tasir, katalog 25 beden.

Kaynak: docs/eon/pricing/2026-07-29-eon-etsy-giris-grid-spot4090-v4.xlsx
        sha256 f198fc27de356ffbf2501bd61d75aba308d6d6c3bd78c7e9bccd13d307f6f3b8

Formuller (429 milgrain + 429 standart hucrede birebir dogrulandi):
    landed_gorunen = ROUND(ham, 2)
    ham            = gram*spot_g*saflik*(1+fire) + iscilik + paket + kargo
    motor          = ROUND(ham * carpan(genislik), 0)     <- HAM landed'den
    etsy_liste     = CEILING(motor * 4/15) * 5            <- Etsy'ye girilen
    gorunen_sale   = etsy_liste * 0.75                    <- kalici %25 indirim
    floor          = ROUND((landed_gorunen + 0.45) / (1 - etsy_fee), 0)
    offsite_floor  = ROUND((landed_gorunen + 0.45) / (1 - etsy_fee - offsite), 0)

Iscilik iki kademeli: standart 30 · MILGRAIN/HAMMERED 40 (ASM: "Hammered/
milgrain istisnasi"). Carpan genislige gore bolunmus: 2-7mm 1.55 · 8-12mm 2.00
("mens wide" primi). Kargo landed'in ICINDE — ayri kargo payi EKLENMEZ.

openpyxl gerekir (build-time script; uygulama bagimliligi degil):
    pip install openpyxl
"""

import hashlib
import math
import pathlib

XLSX = (
    pathlib.Path(__file__).resolve().parent.parent
    / "docs" / "eon" / "pricing"
    / "2026-07-29-eon-etsy-giris-grid-spot4090-v4.xlsx"
)
XLSX_SHA256 = "f198fc27de356ffbf2501bd61d75aba308d6d6c3bd78c7e9bccd13d307f6f3b8"

PROFILE_STANDARD = "standard"
PROFILE_MILGRAIN = "milgrain"


def _round_half_up(x: float, digits: int = 0) -> float:
    """Excel ROUND — yarim yukari (Python'un banker's rounding'i DEGIL)."""
    f = 10 ** digits
    return math.floor(x * f + 0.5) / f


class PricingEngine:
    """v4 grid'in varsayimlari + formulu. Grid'i dogrular, ara bedeni doldurur."""

    def __init__(self) -> None:
        import openpyxl  # yerel import: yalniz bu script'e gerekli

        data = XLSX.read_bytes()
        got = hashlib.sha256(data).hexdigest()
        assert got == XLSX_SHA256, (
            f"v4 grid sha256 uyusmuyor:\n  beklenen {XLSX_SHA256}\n  bulunan  {got}"
        )
        wb = openpyxl.load_workbook(XLSX, data_only=True)
        self._wb = wb
        a = {r[0]: r[1] for r in wb["ASM"].iter_rows(values_only=True) if r[0]}
        self.assumptions = a
        self.spot_per_gram = a["Spot USD/gram"]
        self.spot_per_ozt = a["Spot USD/ozt"]
        self.fire = a["Fire"]
        self.labor = {
            PROFILE_STANDARD: a["Iscilik USD"],
            PROFILE_MILGRAIN: a["Iscilik Milgrain USD"],
        }
        self.packaging = a["Paket USD"]
        self.shipping = a["Kargo USD"]
        self.mult_narrow = a["Carpan 2-7mm"]
        self.mult_wide = a["Carpan 8-12mm"]
        self.etsy_fee = a["Etsy kesinti"]
        self.offsite = a["Offsite Ads"]
        self.purity = {
            "10K": a["Saflik 10K"], "14K": a["Saflik 14K"], "18K": a["Saflik 18K"],
        }
        self.thickness = a["Kalinlik"]
        self._grid = self._load_grid()

    # ── formul ────────────────────────────────────────────────────────
    def multiplier(self, width_mm: float) -> float:
        return self.mult_narrow if width_mm <= 7 else self.mult_wide

    def price(self, karat: str, width_mm: float, grams: float,
              profile: str = PROFILE_MILGRAIN) -> dict:
        raw = (grams * self.spot_per_gram * self.purity[karat] * (1 + self.fire)
               + self.labor[profile] + self.packaging + self.shipping)
        landed = _round_half_up(raw, 2)
        motor = int(_round_half_up(raw * self.multiplier(width_mm)))
        liste = math.ceil(motor * 4 / 15) * 5
        return {
            "landed_cents": int(_round_half_up(landed * 100)),
            "engine_cents": motor * 100,
            "list_cents": liste * 100,
            "sale_cents": int(_round_half_up(liste * 0.75 * 100)),
            "floor_cents": int(_round_half_up((landed + 0.45) / (1 - self.etsy_fee))) * 100,
            "offsite_floor_cents":
                int(_round_half_up((landed + 0.45) / (1 - self.etsy_fee - self.offsite))) * 100,
        }

    # ── grid ──────────────────────────────────────────────────────────
    def _load_grid(self) -> dict:
        """MILGRAIN + 10K/14K/18K sayfalarini (karat,profil,genislik,beden) ile indeksler."""
        grid: dict = {}
        ws = self._wb["MILGRAIN"]
        hdr = [c for c in next(ws.iter_rows(values_only=True))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            d = dict(zip(hdr, row))
            grid[(d["Karat"], PROFILE_MILGRAIN, int(d["Genislik mm"]),
                  float(d["Beden US"]))] = d
        for karat in ("10K", "14K", "18K"):
            ws = self._wb[karat]
            hdr = [c for c in next(ws.iter_rows(values_only=True))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                d = dict(zip(hdr, row))
                grid[(karat, PROFILE_STANDARD, int(d["Genislik mm"]),
                      float(d["Beden US"]))] = d
        return grid

    def grid_grams(self, karat: str, profile: str = PROFILE_MILGRAIN) -> dict[int, dict[float, float]]:
        """Grid'in tam-beden gram tablosu: {genislik: {beden: gram}}."""
        out: dict[int, dict[float, float]] = {}
        for (k, p, w, s), d in self._grid.items():
            if k == karat and p == profile:
                out.setdefault(w, {})[s] = d["Gram 1.5mm"]
        return out

    def grams_with_halves(self, karat: str, profile: str = PROFILE_MILGRAIN
                          ) -> dict[int, dict[float, float]]:
        """Yarim bedenleri komsu tam bedenlerin dogrusal orta-noktasiyla doldurur.

        Ev yontemi (0101/katalog-v3): `(gram[n] + gram[n+1]) / 2`, 3 ondalik.
        Gram tablolari satir-dogrusal oldugu icin interpolasyon fizikseldir.
        """
        base = self.grid_grams(karat, profile)
        out: dict[int, dict[float, float]] = {}
        for width, by_size in base.items():
            sizes = sorted(by_size)
            filled = dict(by_size)
            for lo, hi in zip(sizes, sizes[1:]):
                if hi - lo == 1:
                    filled[lo + 0.5] = round((by_size[lo] + by_size[hi]) / 2, 3)
            out[width] = filled
        return out

    def verify_against_grid(self) -> tuple[int, list]:
        """Grid'in HER hucresini formulle yeniden hesaplar. (kontrol, sapmalar)"""
        checks = 0
        bad = []
        for (karat, profile, width, size), d in self._grid.items():
            got = self.price(karat, width, d["Gram 1.5mm"], profile)
            want = {
                "landed_cents": int(_round_half_up(d["Landed USD"] * 100)),
                "engine_cents": int(d["Motor USD"]) * 100,
                "list_cents": int(d["ETSY LISTE USD"]) * 100,
            }
            if "Gorunen Sale USD" in d and d["Gorunen Sale USD"] is not None:
                want["sale_cents"] = int(_round_half_up(d["Gorunen Sale USD"] * 100))
            if d.get("Floor USD") is not None:
                want["floor_cents"] = int(d["Floor USD"]) * 100
            if d.get("Offsite Floor USD") is not None:
                want["offsite_floor_cents"] = int(d["Offsite Floor USD"]) * 100
            for key, expected in want.items():
                checks += 1
                if got[key] != expected:
                    bad.append((karat, profile, width, size, key, got[key], expected))
        return checks, bad


if __name__ == "__main__":
    e = PricingEngine()
    print(f"v4 grid: {XLSX.name}")
    print(f"  sha256    : OK ({XLSX_SHA256[:16]}…)")
    print(f"  spot      : ${e.spot_per_ozt}/ozt  (${e.spot_per_gram:.4f}/g)")
    print(f"  kalinlik  : {e.thickness}")
    print(f"  iscilik   : standart ${e.labor[PROFILE_STANDARD]} · "
          f"MILGRAIN/HAMMERED ${e.labor[PROFILE_MILGRAIN]}")
    print(f"  carpan    : 2-7mm {e.mult_narrow} · 8-12mm {e.mult_wide}")
    print(f"  paket/kargo: ${e.packaging} / ${e.shipping}  (kargo landed'in ICINDE)")
    print(f"  Etsy/offsite: %{e.etsy_fee*100:.1f} / %{e.offsite*100:.0f}")
    checks, bad = e.verify_against_grid()
    print(f"\nGRID DOGRULAMA: {checks} kontrol -> "
          + ("SIFIR SAPMA" if not bad else f"{len(bad)} SAPMA {bad[:3]}"))
