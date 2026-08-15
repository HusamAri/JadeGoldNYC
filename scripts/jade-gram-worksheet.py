#!/usr/bin/env python3
"""
Jade Gold — gram & fiyat çalışma dosyası (XLSX) üreteci.

NEDEN VAR: 121 varyantın gramı yok ve gram olmadan maliyet, dolayısıyla marj
hesaplanamıyor. ShipStation kaynağı kurudu (gram oraya *internal notes* alanına
elle yazılıyor ve notu dolu olan tüm SKU'lar zaten çekilmiş), açıklama metni
yalnız tek-varyantlı listinglerde güvenilir. Kalan tek yol terazi + elle giriş;
bu dosya ekibin o girişi yapacağı yer.

FORMÜLLER HÜCREDE CANLI: ekip gram yazdığı anda maliyet/hedef fiyat/marj kendi
hesaplanır. Tüm sabitler `Ayarlar` sekmesinde tek yerde — altın fiyatı değişince
bir hücre güncellenir, 1.944 satır yeniden hesaplanır.

Formüller `lib/jade/pricing.ts` ile AYNI olmalı; ikisi ayrı yerde yaşıyor, biri
değişirse diğeri de değişmeli (SQL↔TS drift dersi).

Kullanım:  python3 scripts/jade-gram-worksheet.py catalog.json cikti.xlsx
catalog.json biçimi: [{"ch","lid","title","sku","g","pc","ws"}, ...]
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Jade Gold marka paleti ────────────────────────────────────────────────
# Kaynak: public/brand/jade-gold-nyc-guidelines.html (dosyadaki baskın renkler).
# Panelin globals.css'i DEĞİL — orası mor tonlu bir arayüz teması, marka değil.
ALTIN = "9A7A33"        # ana marka altını
ALTIN_ACIK = "B89347"
ALTIN_KOYU = "7C5F25"
KREM = "F2EFE6"         # arka plan / zebra
KREM_ACIK = "FBF9F4"
KOYU = "1C1A16"         # metin
JADE = "3F4A44"         # jade yeşili — "sağlıklı" sinyali
BEJ = "A39F94"          # nötr / veri yok
GIRIS = "FCEFC7"        # doldurulacak hücre — altının açık tonu, krem'den ayrışır
UYARI = "B00020"        # zarar/şüphe — palet dışı ama uyarı okunaklı olmalı

# Ayarlar sekmesindeki hücre adresleri — formüller buraya bağlanır.
# DİKKAT: sekmede 1. satır başlık bandı, 2. satır kolon adları → veri 3'ten başlar.
# Bu adresler kayarsa formüller sessizce yanlış hücreye bakar.
C_GRAM = "Ayarlar!$B$3"      # $/gram (her şey dahil tedarikçi)
C_SHIP = "Ayarlar!$B$4"      # kargo
C_PACK = "Ayarlar!$B$5"      # ambalaj
C_DISC = "Ayarlar!$B$6"      # kalıcı indirim
C_RATE = "Ayarlar!$B$7"      # Etsy oransal kesinti
C_FIX = "Ayarlar!$B$8"       # Etsy sabit kesinti
C_MARGIN = "Ayarlar!$B$9"    # hedef marj

F_GIRIS = PatternFill("solid", fgColor=GIRIS)
F_BASLIK = PatternFill("solid", fgColor=ALTIN)
F_KREM = PatternFill("solid", fgColor=KREM)
F_KREM_ACIK = PatternFill("solid", fgColor=KREM_ACIK)
F_JADE = PatternFill("solid", fgColor="E4EBE7")
F_UYARI = PatternFill("solid", fgColor="F7DCE0")
F_BEJ = PatternFill("solid", fgColor="EDEBE6")

BASLIK = Font(bold=True, color="FFFFFF", size=11)
GOVDE = Font(color=KOYU, size=10)
VURGU = Font(bold=True, color=ALTIN_KOYU, size=10)
KIRMIZI = Font(color=UYARI, bold=True, size=10)

_ince = Side(style="thin", color=BEJ)
CERCEVE = Border(bottom=_ince)

# Fiyatı bunun üstünde olan satır placeholder sayılır (canlıda $34.999,99 gibi
# dolgu değerler var; gram girilse bile marjları anlamsız çıkar).
SUPHELI_FIYAT = 30000


def ayarlar_sekmesi(wb):
    ws = wb.create_sheet("Ayarlar", 0)
    ws.sheet_properties.tabColor = ALTIN
    ws["A1"] = "JADE GOLD NYC — FİYAT AYARLARI"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = F_BASLIK
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws["A2"], ws["B2"], ws["C2"] = "AYAR", "DEĞER", "AÇIKLAMA"
    for h in ("A2", "B2", "C2"):
        ws[h].font = Font(bold=True, color=ALTIN_KOYU, size=10)
        ws[h].fill = F_KREM
    satirlar = [
        ("Altın maliyeti ($/gram)", 106, "Her şey dahil tedarikçi faturası: ham altın + işçilik + taş"),
        ("Kargo ($)", 5, "Ücretsiz kargo — bedeli satıcıda"),
        ("Ambalaj ($)", 1.5, ""),
        ("Etsy indirimi", 0.15, "Kalıcı mağaza indirimi — tahsilat = liste × 0,85"),
        ("Etsy oransal kesinti", 0.095, "%6,5 işlem + %3 ödeme"),
        ("Etsy sabit kesinti ($)", 0.25, ""),
        ("Hedef net marj", 0.20, "Fiyat önerisi bu marja göre hesaplanır"),
    ]
    for i, (ad, deger, aciklama) in enumerate(satirlar, start=3):
        ws.cell(i, 1, ad).font = GOVDE
        c = ws.cell(i, 2, deger)
        c.font = VURGU
        c.fill = F_GIRIS          # bu hücreler düzenlenebilir
        c.alignment = Alignment(horizontal="center")
        ws.cell(i, 3, aciklama).font = Font(color=BEJ, size=9, italic=True)
        for col in range(1, 4):
            ws.cell(i, col).border = CERCEVE
    for r in (6, 7, 9):           # yüzde biçimli satırlar (indirim/oran/marj)
        ws.cell(r, 2).number_format = "0.0%"
    for r in (3, 4, 5, 8):
        ws.cell(r, 2).number_format = '"$"#,##0.00'
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 64
    ws["A11"] = "Sarı hücreler düzenlenebilir. Değiştirince tüm sekmeler yeniden hesaplanır."
    ws["A11"].font = Font(italic=True, color=ALTIN_KOYU, size=10)
    ws.merge_cells("A11:C11")
    ws.sheet_view.showGridLines = False
    return ws


def _formuller(r, gram_h, fiyat_h):
    """Bir satırın hesap formülleri. gram_h/fiyat_h = sütun harfleri."""
    g, f = f"{gram_h}{r}", f"{fiyat_h}{r}"
    maliyet = f"{g}*{C_GRAM}+{C_SHIP}+{C_PACK}"
    tahsilat = f"{f}*(1-{C_DISC})"
    net = f"{tahsilat}-({tahsilat}*{C_RATE}+{C_FIX})-({maliyet})"
    return {
        "maliyet": f'=IF({g}="","",{maliyet})',
        "breakeven": f'=IF({g}="","",ROUNDUP((({maliyet})+{C_FIX})/((1-{C_DISC})*(1-{C_RATE})),0))',
        "hedef": f'=IF({g}="","",ROUNDUP((({maliyet})+{C_FIX})/((1-{C_DISC})*(1-{C_RATE}-{C_MARGIN})),0))',
        "marj": f'=IF(OR({g}="",{f}=""),"",({net})/({tahsilat}))',
    }


def _durum(r, gram_h, fiyat_h, marj_h, breakeven_h):
    g, f = f"{gram_h}{r}", f"{fiyat_h}{r}"
    return (
        f'=IF({g}="","GRAM EKSİK",'
        f'IF({f}>{SUPHELI_FIYAT},"FİYAT ŞÜPHELİ",'
        f'IF({marj_h}{r}<0,"ZARARDA",'
        f'IF({f}<{breakeven_h}{r},"RİSKLİ","ok"))))'
    )


def veri_sekmesi(wb, ad, rows, kaynak_sutunu, neden_sutunu=False):
    ws = wb.create_sheet(ad)
    basliklar = ["Bölüm", "Listing", "Etsy ID", "SKU"]
    if neden_sutunu:
        basliklar.append("Neden")
    if kaynak_sutunu:
        basliklar.append("Gram Kaynağı")
    basliklar += ["Fiyat $", "GRAM", "Maliyet $", "Breakeven $",
                  "Hedef Fiyat $", "Marj", "Durum", "Etsy Linki"]
    ws.sheet_properties.tabColor = ALTIN_ACIK if kaynak_sutunu else ALTIN
    ws.append(basliklar)
    for c in range(1, len(basliklar) + 1):
        ws.cell(1, c).font = BASLIK
        ws.cell(1, c).fill = F_BASLIK
        ws.cell(1, c).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Sütun harfleri başlık sırasına göre çözülür (iki sekmede farklı).
    idx = {h: get_column_letter(i + 1) for i, h in enumerate(basliklar)}
    F, G = idx["Fiyat $"], idx["GRAM"]
    K, M, H, B = idx["Marj"], idx["Maliyet $"], idx["Hedef Fiyat $"], idx["Breakeven $"]

    for i, r in enumerate(rows, start=2):
        satir = [r["ch"], r["title"], r["lid"], r["sku"]]
        if neden_sutunu:
            satir.append(r.get("rc") or "gram yok")
        if kaynak_sutunu:
            satir.append(r["ws"] or "—")
        satir += [
            (r["pc"] / 100) if r["pc"] is not None else None,
            r["g"],  # None ise boş kalır → ekip doldurur
        ]
        ws.append(satir)
        fm = _formuller(i, G, F)
        ws[f"{M}{i}"] = fm["maliyet"]
        ws[f"{B}{i}"] = fm["breakeven"]
        ws[f"{H}{i}"] = fm["hedef"]
        ws[f"{K}{i}"] = fm["marj"]
        ws[f"{idx['Durum']}{i}"] = _durum(i, G, F, K, B)
        ws[f"{idx['Etsy Linki']}{i}"] = f'https://www.etsy.com/listing/{r["lid"]}'
        # Zebra: krem/beyaz almaşık — 1.944 satırda gözün satırı kaybetmemesi için.
        if i % 2 == 0:
            for c in range(1, len(basliklar) + 1):
                ws.cell(i, c).fill = F_KREM_ACIK
        for c in range(1, len(basliklar) + 1):
            ws.cell(i, c).font = GOVDE
            ws.cell(i, c).border = CERCEVE
        # Sarı giriş alanı: gramı olmayan VE tartımı tartışmalı olanlar.
        # rc = gramı var ama güvenilmiyor (bkz. 1739245557 beden sorunu) —
        # mevcut değer görünür kalır ki ekip neyi düzelttiğini bilsin.
        if r["g"] is None or r.get("rc"):
            ws[f"{G}{i}"].fill = F_GIRIS      # zebra'yı EZER: giriş alanı hep görünür
            ws[f"{G}{i}"].font = VURGU
        ws[f"{G}{i}"].alignment = Alignment(horizontal="center")
        for col in (F, M, B, H):
            ws[f"{col}{i}"].number_format = '"$"#,##0.00'
        ws[f"{K}{i}"].number_format = "0.0%"
        ws[f"{idx['Etsy Linki']}{i}"].font = Font(color="0563C1", underline="single", size=9)

    # Koşullu biçim: durum METNİNE göre renk. Statik değil — ekip gram yazınca
    # formül yeniden hesaplanır ve renk kendiliğinden değişir.
    son = len(rows) + 1
    D = idx["Durum"]
    kurallar = [
        ("ZARARDA", F_UYARI, KIRMIZI),
        ("FİYAT ŞÜPHELİ", F_UYARI, KIRMIZI),
        ("RİSKLİ", PatternFill("solid", fgColor="FBEFD5"), Font(color=ALTIN_KOYU, bold=True, size=10)),
        ("ok", F_JADE, Font(color=JADE, bold=True, size=10)),
        ("GRAM EKSİK", F_BEJ, Font(color=BEJ, italic=True, size=10)),
    ]
    for metin, dolgu, yazi in kurallar:
        ws.conditional_formatting.add(
            f"{D}2:{D}{son}",
            CellIsRule(operator="equal", formula=[f'"{metin}"'], fill=dolgu, font=yazi),
        )
    # Marj sütunu: kırmızı → krem → jade renk skalası (sayısal okumayı hızlandırır)
    ws.conditional_formatting.add(
        f"{K}2:{K}{son}",
        ColorScaleRule(start_type="num", start_value=-0.3, start_color="F5C6CB",
                       mid_type="num", mid_value=0.10, mid_color=KREM,
                       end_type="num", end_value=0.40, end_color="CFE3D7"),
    )

    genislik = {"Bölüm": 12, "Listing": 46, "Etsy ID": 12, "SKU": 16,
                "Neden": 26, "Gram Kaynağı": 14, "Fiyat $": 11, "GRAM": 9,
                "Maliyet $": 11, "Breakeven $": 12, "Hedef Fiyat $": 13,
                "Marj": 9, "Durum": 15, "Etsy Linki": 42}
    for h, harf in idx.items():
        ws.column_dimensions[harf].width = genislik.get(h, 12)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(basliklar))}{len(rows) + 1}"
    ws.sheet_view.showGridLines = False
    return ws


def ozet_sekmesi(wb, rows):
    ws = wb.create_sheet("Listing Özeti")
    ws.sheet_properties.tabColor = JADE
    ws.append(["Bölüm", "Listing", "Etsy ID", "Varyant", "Gramsız",
               "Min Fiyat $", "Max Fiyat $", "Durum", "Etsy Linki"])
    for c in range(1, 10):
        ws.cell(1, c).font = BASLIK
        ws.cell(1, c).fill = F_BASLIK
        ws.cell(1, c).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    gruplar = {}
    for r in rows:
        gruplar.setdefault(r["lid"], []).append(r)

    sirali = sorted(gruplar.items(),
                    key=lambda kv: (-sum(1 for x in kv[1] if x["g"] is None), kv[1][0]["ch"]))
    for lid, grup in sirali:
        fiyatlar = [x["pc"] / 100 for x in grup if x["pc"] is not None]
        eksik = sum(1 for x in grup if x["g"] is None)
        ws.append([
            grup[0]["ch"], grup[0]["title"], lid, len(grup), eksik,
            min(fiyatlar) if fiyatlar else None,
            max(fiyatlar) if fiyatlar else None,
            "TÜMÜ EKSİK" if eksik == len(grup) else ("kısmen eksik" if eksik else "tam"),
            f"https://www.etsy.com/listing/{lid}",
        ])
    for i in range(2, ws.max_row + 1):
        if i % 2 == 0:
            for c in range(1, 10):
                ws.cell(i, c).fill = F_KREM_ACIK
        for c in range(1, 10):
            ws.cell(i, c).font = GOVDE
            ws.cell(i, c).border = CERCEVE
        if ws.cell(i, 5).value:
            ws.cell(i, 5).font = KIRMIZI
        for c in (4, 5):
            ws.cell(i, c).alignment = Alignment(horizontal="center")
        for c in (6, 7):
            ws.cell(i, c).number_format = '"$"#,##0.00'
        ws.cell(i, 9).font = Font(color="0563C1", underline="single", size=9)

    son = ws.max_row
    for metin, dolgu, yazi in (
        ("TÜMÜ EKSİK", F_UYARI, KIRMIZI),
        ("kısmen eksik", PatternFill("solid", fgColor="FBEFD5"),
         Font(color=ALTIN_KOYU, bold=True, size=10)),
        ("tam", F_JADE, Font(color=JADE, bold=True, size=10)),
    ):
        ws.conditional_formatting.add(
            f"H2:H{son}",
            CellIsRule(operator="equal", formula=[f'"{metin}"'], fill=dolgu, font=yazi),
        )

    for harf, w in zip("ABCDEFGHI", (12, 46, 12, 9, 9, 12, 12, 14, 42)):
        ws.column_dimensions[harf].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{son}"
    ws.sheet_view.showGridLines = False
    return ws


def main():
    kaynak, cikti = sys.argv[1], sys.argv[2]
    rows = json.load(open(kaynak, encoding="utf-8"))
    # Ekip girişi iki sebeple istenir: gram HİÇ yok, ya da gram var ama
    # güvenilmiyor (rc = recheck; 1739245557'de 10 beden tek gramı paylaşıyor).
    giris = [r for r in rows if r["g"] is None or r.get("rc")]

    wb = Workbook()
    wb.remove(wb.active)
    ayarlar_sekmesi(wb)
    veri_sekmesi(wb, "Gram Girişi", giris, kaynak_sutunu=False, neden_sutunu=True)
    veri_sekmesi(wb, "Tüm Varyantlar", rows, kaynak_sutunu=True)
    ozet_sekmesi(wb, rows)
    wb.save(cikti)

    yok = sum(1 for r in giris if r["g"] is None)
    print(f"{cikti} yazıldı")
    print(f"  Gram Girişi   : {len(giris)} satır ({yok} gram yok, {len(giris)-yok} tartım gerekli)")
    print(f"  Tüm Varyantlar: {len(rows)} satır")
    print(f"  Listing Özeti : {len(set(r['lid'] for r in rows))} listing")


if __name__ == "__main__":
    main()
