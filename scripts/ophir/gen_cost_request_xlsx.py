# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
YELLOW = PatternFill("solid", fgColor="FFFF00")
HDRFILL = PatternFill("solid", fgColor="1F3864")
FILLHDR = PatternFill("solid", fgColor="C00000")
REFFILL = PatternFill("solid", fgColor="F2F2F2")
WARNFILL = PatternFill("solid", fgColor="FFC7CE")
BLUE = Font(name=FONT, color="0000FF", size=10)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# (n, etsy_id, title, status, width, shank, thick, title_mm, stone, form, est_g14, url_slug)
D = [
(1,4556227760,"11mm Dome Band Ring | Chunky Minimalist Statement Ring, Stacking Jewelry","edit",11,3.5,1.5,11,"yok / none","dome",5.13,"4556227760/dome-band-ring-wide-dome-ring-chunky"),
(2,4543752254,"1mm Thin Stacking Ring | Minimalist Dainty Band | Simple Everyday Gold Silver Jewelry","edit",1.5,None,1.5,1,"tekil/aksan tas","band",1.48,"4543752254/1mm-thin-stacking-ring-minimalist-dainty"),
(3,4546852268,"2mm Smooth Gold Band Ring | Minimalist Wedding, Stacking, Promise Jewelry","edit",2,None,1.5,2,"tam eternity","band",1.98,"4546852268/2mm-smooth-gold-band-ring-minimalist"),
(4,4549686009,"4mm Curved Wedding Band Ring | Minimalist Stacking Bridal Ring","edit",4,None,1.5,4,"yok / none","band",3.96,"4549686009/4mm-curved-wedding-band-ring-minimalist"),
(5,4543147022,"4mm Smooth Band Ring | Minimalist Wedding Stacking Ring","edit",4,None,1.5,4,"yok / none","band",3.96,"4543147022/4mm-smooth-band-ring-minimalist-wedding"),
(6,4547299090,"4mm Summit Band Ring | Minimalist Wedding, Stacking, Promise Band | Modern Comfort Fit","edit",4,4,1.5,4,"yok / none","band",3.96,"4547299090/4mm-summit-band-ring-minimalist-wedding"),
(7,4543746594,"6mm Curved Band Ring | Minimalist Gold Silver Modern Stacking Jewelry","edit",6,None,1.5,6,"yok / none","band",5.93,"4543746594/6mm-curved-band-ring-minimalist-gold"),
(8,4543233648,"6mm Smooth Wedding Band | Minimalist Comfort Fit Ring, Stackable Gift","edit",1.5,None,1.5,6,"yok / none","band",1.48,"4543233648/6mm-smooth-wedding-band-minimalist"),
(9,4558001828,"7mm Pave Diamond Star Open Ring | Celestial Minimalist Jewelry","edit",7,None,1.5,7,"pave","band",6.92,"4558001828/7mm-pave-diamond-star-open-ring-o"),
(10,4552138588,"Aurora Lab Grown Sapphire Eternity Ring | Blue Sapphire Band, 3.10mm Width","edit",1.5,None,1.5,3.10,"tam eternity","band",1.48,"4552138588/aurora-lab-grown-sapphire-eternity-ring"),
(11,4554955252,"Aurora Oval Eternity Ring | Minimalist Diamond Band, Stackable Wedding Ring","edit",2.5,2.5,1.5,None,"tam eternity","band",2.47,"4554955252/oval-eternity-ring-aurora-diamond-ring"),
(12,4558666025,"Baguette Diamond Eternity Ring | Thin Stackable Wedding Band | Minimalist Bridal Jewelry","edit",1.95,None,1.5,None,"tam eternity (baget)","band",1.93,"4558666025/baguette-eternity-ring-diamond-wedding"),
(13,4559889337,"Baguette Diamond Half Eternity Ring | 1.45mm Slim Stacking Wedding Band","edit",1.45,None,1.5,1.45,"yarim eternity (baget)","band",1.43,"4559889337/baguette-diamond-half-eternity-ring"),
(14,4559901602,"Baguette Emerald Half Eternity Ring | Minimalist Stackable Band (2.35mm Top)","edit",2.35,1.15,1.5,2.35,"yarim eternity (baget)","band",1.55,"4559901602/emerald-baguette-half-eternity-ring"),
(15,4559862879,"Baguette Emerald Open Ring | Minimalist Gold Gemstone Stackable Band","edit",1.1,1.1,1.5,None,"baget","band",1.09,"4559862879/baguette-emerald-open-ring-emerald-green"),
(16,4553815586,"Baguette Ring | Minimalist Stackable Diamond Band | Everyday Wedding Ring","edit",1,1,1.5,1,"baget","band",0.99,"4553815586/baguette-ring-minimalist-stackable"),
(17,4556230955,"Bezel Emerald Half Eternity Band | Minimalist Stacking Ring | Anniversary Gift","edit",2.5,1.3,1.5,None,"yarim eternity","band",1.70,"4556230955/bezel-emerald-half-eternity-band-emerald"),
(18,4556240933,"Boa Gemstone Coil Ring | Serpent Inspired Wrap Statement Ring","edit",9.7,None,1.5,None,"tekil/aksan tas","band",9.59,"4556240933/boa-gemstone-coil-ring-serpent-inspired"),
(19,4549692483,"Bold Diamond Eternity Band | 4mm Wedding Ring, Minimalist Bridal Stacking Band","edit",4,None,1.5,4,"tam eternity","band",3.96,"4549692483/bold-diamond-eternity-band-4mm-wedding"),
(20,4544917788,"Bold Diamond Eternity Ring | 2.6mm Gold Wedding Band, Stacking Ring","edit",2.6,None,1.5,2.6,"tam eternity","band",2.57,"4544917788/bold-diamond-eternity-ring-26mm-gold"),
(21,4552128868,"Bold Dot Ring | Minimalist Stacking Band | Modern Everyday Jewelry (3.6mm Wide)","edit",1.6,None,1.6,3.6,"yok / none","band",1.70,"4552128868/bold-dot-ring-minimalist-stacking-band"),
(22,4545510477,"Bold Line Eternity Band Ring | 3mm Minimalist Stacking Wedding Anniversary Band","edit",3,None,1.5,3,"tam eternity","band",2.97,"4545510477/bold-line-eternity-band-ring-3mm"),
(23,4550930017,"Bold Puzzle Stacking Ring | 3mm Gold Band, Modern Minimalist Statement Jewelry","edit",3,3,1.5,3,"yok / none","band",2.97,"4550930017/bold-puzzle-stacking-ring-3mm-gold-band"),
(24,4551550234,"Bold Round Signet Ring | 13mm Top Width, 4mm Shank | Modern Statement Jewelry","edit",13,4,1.5,13,"yok / none","signet",7.07,"4551550234/bold-round-signet-ring-13mm-top-width"),
(25,4554965833,"Bold Stacker Ring | Minimalist Band, 3mm Width, Everyday Jewelry","edit",3,None,1.5,3,"yok / none","band",2.97,"4554965833/bold-stacking-ring-chunky-stacker-ring"),
(26,4546827813,"Braided Wedding Band | Minimalist Woven Ring, Stackable Jewelry (4.45mm)","edit",4.45,None,1.5,4.45,"tam eternity","band",4.40,"4546827813/braided-wedding-band-minimalist-woven"),
(27,4554957582,"Celestial Dome Ring | Sterling Silver Minimalist Band | Everyday Stackable Jewelry","edit",5.6,1.65,1.5,None,"yok / none","dome",2.54,"4554957582/celestial-dome-ring-sterling-silver-dome"),
(28,4549031600,"Charlotte Bold Gold Ring | Wide Band Minimalist Statement Jewelry","edit",9.35,None,1.5,None,"yok / none","band",9.25,"4549031600/charlotte-bold-gold-ring-wide-band"),
(29,4552100498,"Charlotte Crest Gold Stacking Ring | Minimalist Dainty Band, 2.8mm Everyday Jewelry","edit",2.80,None,1.5,2.8,"yok / none","band",2.77,"4552100498/charlotte-crest-gold-stacking-ring"),
(30,4554964031,"Charlotte Ring | 6.35mm Wide Chunky Band Statement Ring","edit",6.35,None,1.5,6.35,"yok / none","band",6.28,"4554964031/charlotte-ring-chunky-band-ring"),
(31,4550953712,"Charlotte Stacker Ring Set | Minimalist Band, 6.20mm Width, Dainty Jewelry","edit",6.20,None,1.5,6.20,"yok / none","band",6.13,"4550953712/charlotte-stacker-ring-set-minimalist"),
(32,4553157256,"Chunky Dome Ring | Modern Gold Statement Band | Minimalist Jewelry (6.3mm Top Width)","edit",6.3,2,1.5,6.3,"yok / none","dome",2.94,"4553157256/chunky-dome-ring-modern-gold-statement"),
(33,4556249030,"Chunky Dome Statement Ring | Minimalist Gold Band, Stackable Modern Jewelry","edit",5.8,None,1.5,None,"yok / none","dome",4.86,"4556249030/dome-figure-ring-o-dome-statement-ring-o"),
(34,4552026081,"Colosseum Wide Band Ring | Modern Minimalist Gold Statement Jewelry","edit",1.5,None,1.5,None,"yok / none","band",1.48,"4552026081/colosseum-wide-band-ring-modern"),
(35,4552085145,"Courage Signet Ring | Bold Minimalist Statement Jewelry (13mm Top)","edit",13,1.7,1.5,13,"yok / none","signet",5.59,"4552085145/courage-signet-ring-bold-minimalist"),
(36,4551576045,"Curved Gold Band Ring | Minimalist 2mm Stacking Wedding Ring","edit",2,None,1.5,2,"yok / none","band",1.98,"4551576045/curved-gold-band-ring-minimalist-2mm"),
(37,4538876982,"Curved Pave Diamond Nesting Ring, Solid Gold","edit",1.5,None,1.5,None,"pave","band",1.48,"4538876982/dome-pave-diamond-ring-curved-wave"),
(38,4557310374,"Dainty Pearl and Diamond Ring | Slim 1mm Band, Minimalist Jewelry","edit",1,None,1.5,1,"yok / none","band",0.99,"4557310374/beverly-ring-dainty-stacking-ring-slim"),
(39,4545535248,"Diamond Cluster Engagement Ring | Elegant Statement Bridal Wedding Band","edit",7,2,1.5,None,"tekil/aksan tas","band",3.71,"4545535248/diamond-cluster-engagement-ring-elegant"),
(40,4558673647,"Diamond Coil Ring | Modern Spiral Band, 9.8mm Width","edit",9.8,None,1.5,9.8,"tekil/aksan tas","band",9.69,"4558673647/diamond-coil-ring-modern-spiral-band"),
(41,4557295667,"Diamond Eternity Band | Minimalist 2.8mm Wedding Ring for Women","edit",2.8,None,1.5,2.8,"tam eternity","band",2.77,"4557295667/diamond-eternity-band-dainty-diamond"),
(42,4552114869,"Diamond Half Eternity Ring | 2.5mm Bold Wedding Band, Stackable Bridal Jewelry","edit",1.5,None,1.5,2.5,"yarim eternity","band",1.48,"4552114869/diamond-half-eternity-ring-25mm-bold"),
(43,4547301260,"Diamond Stacking Ring | Minimalist Wedding Band, Dainty Promise Ring, Everyday Wear","edit",1.6,1.6,1.5,None,"tekil/aksan tas","band",1.58,"4547301260/diamond-stacking-ring-minimalist-wedding"),
(44,4552097960,"Dome Figure Balance Ring | Sculptural Minimalist Statement (3.4mm Wide)","edit",3.40,None,1.5,3.4,"yok / none","dome",2.85,"4552097960/dome-figure-balance-ring-o-sculptural"),
(45,4544906099,"Duo Beaded Stacker Ring | Minimalist Gold Band, Everyday Jewelry (2.9mm Width)","edit",2.9,None,1.5,2.9,"yok / none","band",2.87,"4544906099/duo-beaded-stacker-ring-minimalist-gold"),
(46,4554958199,"Emerald Diamond Wishbone Ring | Curved Stacking Wedding Band","edit",1,1,1.5,None,"tekil/aksan tas","band",0.99,"4554958199/emerald-diamond-ring-wishbone-wedding"),
(47,4553159638,"Enamel Charlotte Stacker Ring | Colorful Band with 1.75mm Stone, Everyday Minimalist","edit",3.30,None,1.5,1.75,"tekil/aksan tas","band",3.26,"4553159638/enamel-charlotte-stacker-ring-colorful"),
(48,4549712730,"Gemstone Trillion Stacker Ring | Minimalist Wedding Band, 3mm Band","edit",3,None,1.5,3,"tam eternity","band",2.97,"4549712730/gemstone-trillion-stacker-ring"),
(49,4550924858,"Gold Beaded Stacking Ring | Minimalist Everyday Band, Delicate Gift","edit",1.3,None,1.5,None,"yok / none","band",1.29,"4550924858/gold-beaded-stacking-ring-minimalist"),
(50,4557987114,"Gold Dot Band Ring | Minimalist Stackable Everyday Jewelry","edit",2.3,None,1.5,None,"yok / none","band",2.27,"4557987114/dot-band-ring-stackable-ring-minimalist"),
(51,4550284137,"Gold Pave Diamond Signet Ring | Minimalist Statement Pinky Ring","edit",8,1.5,1.5,None,"pave","signet",3.73,"4550284137/gold-pave-diamond-signet-ring-o"),
(52,4550936365,"Gold Pave Puzzle Ring | Minimalist Stacking Band, Everyday Jewelry","edit",1.5,1.5,1.5,None,"pave","band",1.48,"4550936365/gold-pave-puzzle-ring-o-minimalist"),
(53,4550299061,"Gold Signet Ring | Minimalist Pinky Ring, Dainty Stackable Jewelry","edit",7,1.8,1.5,None,"yok / none","signet",3.58,"4550299061/gold-signet-ring-minimalist-pinky-ring"),
(54,4554955373,"Heirloom Dome Ring | Modern Statement Wide Band, 8mm Top, Minimalist Everyday Jewelry","edit",8,2.5,1.5,8,"yok / none","band",4.38,"4554955373/heirloom-ring-dome-ring-modern-statement"),
(55,4556245256,"Interconnected Wave Pave Diamond Ring | Modern Sculptural Statement Jewelry","edit",9.3,1.25,1.5,None,"pave","band",4.02,"4556245256/wave-pave-diamond-ring-o-interlocking"),
(56,4546842479,"Lab Grown Diamond Eternity Band | 1.7mm Wedding Ring, Stackable Anniversary Gift","edit",1.7,1.7,1.5,1.7,"tam eternity","band",1.68,"4546842479/lab-grown-diamond-eternity-band-17mm"),
(57,4552078311,"Lab Grown Diamond Eternity Band | 3.1mm Wedding Ring, Stackable Anniversary Gift","edit",1.5,None,1.5,3.1,"tam eternity","band",1.48,"4552078311/lab-grown-diamond-eternity-band-31mm"),
(58,4552097077,"Lab Grown Diamond Line Ring | Minimalist Stacking Band (1.3mm Width)","edit",1.5,None,1.5,1.3,"tekil/aksan tas","band",1.48,"4552097077/lab-grown-diamond-line-ring-minimalist"),
(59,4553792547,"Lab Grown Sapphire Dome Ring | Pave Statement Band, 6.5mm Width","edit",6.5,None,1.5,6.5,"pave","dome",5.45,"4553792547/lab-grown-sapphire-dome-ring-o-pave"),
(60,4557989576,"Lacey Lab Grown Diamond Band | Delicate Stacking Wedding Ring","edit",2.3,None,1.5,None,"tekil/aksan tas","band",2.27,"4557989576/lab-grown-diamond-ring-lacey-diamond"),
(61,4553155082,"Lacey Lab Grown Sapphire Ring | Dainty Engagement Ring | Minimalist Gemstone Jewelry","edit",2.5,1.5,1.5,2.5,"tekil/aksan tas","band",1.83,"4553155082/lacey-lab-grown-sapphire-ring-dainty"),
(62,4546179727,"Meridian Band Ring | Minimalist Dome Design, Unisex Contemporary Jewelry","edit",6,6,1.5,None,"yok / none","dome",5.03,"4546179727/meridian-band-ring-minimalist-dome"),
(63,4553808001,"Metallic Sphere Ring | Minimalist 3mm Wide Band | Everyday Stacking Jewelry","edit",3,None,1.5,3,"yok / none","band",2.97,"4553808001/metallic-sphere-ring-minimalist-3mm-wide"),
(64,4546184741,"Minimalist Gold Dome Ring | Modern Open Band Everyday Jewelry (5mm Wide)","edit",5,1.6,1.5,5,"yok / none","dome",2.34,"4546184741/minimalist-gold-dome-ring-modern-open"),
(65,4549022637,"Minimalist Gold Dome Ring | Slim Stacking Band | Everyday Statement Jewelry","edit",3,1.6,1.5,None,"yok / none","dome",1.75,"4549022637/minimalist-gold-dome-ring-slim-stacking"),
(66,4551533957,"Minimalist Gold Eternity Band Ring | Stackable Wedding Anniversary Band","edit",2,None,1.5,None,"tam eternity","band",1.98,"4551533957/minimalist-gold-eternity-band-ring"),
(67,4550934548,"Minimalist Gold Signet Ring | Classic 6mm Pinky Band Jewelry","edit",6,1,1.5,6,"yok / none","signet",2.72,"4550934548/minimalist-gold-signet-ring-classic-6mm"),
(68,4552101181,"Minimalist Gold Stacker Ring | 5.8mm Wide Band, Everyday Wear","edit",5.8,None,1.5,5.8,"yok / none","band",5.74,"4552101181/minimalist-gold-stacker-ring-58mm-wide"),
(69,4552033260,"Minimalist Gold Wishbone Ring | Dainty V Stacking Band, Everyday Wear","edit",1,None,1.5,None,"yok / none","band",0.99,"4552033260/minimalist-gold-wishbone-ring-dainty-v"),
(70,4554940895,"Open Diamond Band Ring | Minimalist Stackable Thin Ring","edit",1.2,None,1.5,None,"tekil/aksan tas","band",1.19,"4554940895/open-diamond-ring-delicate-diamond-band"),
(71,4559884744,"Open Dome Pave Diamond Ring | Minimalist Open Band Jewelry","edit",5.5,1.7,1.5,None,"pave","dome",2.54,"4559884744/pave-diamond-ring-open-dome-ring-open"),
(72,4558660905,"Organic Dome Marquise Ring | Thin Diamond Stacking Band, Minimalist Bridal Nesting Ring","edit",3.80,1.7,1.5,None,"tekil/aksan tas","dome",2.04,"4558660905/marquise-cut-nesting-ring-organic-dome"),
(73,4552133042,"Organic Dome Nesting Ring Set | Minimalist Stackable Rings (1.5mm)","edit",1.5,None,1.5,1.5,"yok / none","dome",1.26,"4552133042/organic-dome-nesting-ring-set-o"),
(74,4559892713,"Pave Diamond X Ring | Modern Stackable Statement Band (9.5mm Wide)","edit",9.5,None,1.5,9.5,"pave","band",9.40,"4559892713/pave-diamond-x-ring-o-x-ring-band-o"),
(75,4559891185,"Prism Lab Grown Diamond Eternity Band | Geometric Wedding Stacking Ring 2.2mm","edit",2.2,2.2,1.5,2.2,"tam eternity","band",2.18,"4559891185/lab-grown-diamond-eternity-band-prism"),
(76,4554965144,"Puffy Charlotte Dome Ring | 6.2mm Wide Stacker Band, Minimalist Statement Jewelry","edit",6.20,None,1.5,6.2,"yok / none","dome",5.19,"4554965144/puffy-charlotte-ring-chunky-stacker-ring"),
(77,4554953887,"Puzzle Stacking Ring | 1.5mm Thin Minimalist Band | Modern Everyday Jewelry","edit",1.5,None,1.5,1.5,"yok / none","band",1.48,"4554953887/puzzle-ring-stacking-ring-minimalist"),
(78,4550939147,"Sculptural Band Ring | Modern Minimalist, 2.5mm Width, Stackable Everyday Jewelry","edit",2.5,None,1.5,2.5,"yok / none","band",2.47,"4550939147/sculptural-band-ring-modern-minimalist"),
(79,4557297951,"Sculptural Open Dome Ring | Bold Minimalist Band, Contemporary Statement Jewelry","edit",10.6,2,1.5,None,"yok / none","dome",4.20,"4557297951/open-dome-ring-bold-dome-ring-sculptural"),
(80,4558664005,"Seismic Band Ring | 6mm Wide Contemporary Wedding Band, Minimalist Stackable Jewelry","edit",6,6,1.5,6,"yok / none","band",5.93,"4558664005/seismic-ring-band-6mm-chunky-band-ring"),
(81,4559909674,"Slim Diamond Eternity Band | 1.5mm Minimalist Wedding Ring","edit",1.5,None,1.5,1.5,"tam eternity","band",1.48,"4559909674/diamond-eternity-band-dainty-wedding"),
(82,4559881595,"Slim Diamond Eternity Band | Dainty Stacking Ring (1.1mm)","edit",1.1,None,1.5,1.1,"tam eternity","band",1.09,"4559881595/diamond-eternity-band-slim-diamond-ring"),
(83,4558668281,"Slim Gold Signet Ring | Dainty Minimalist Stacking Band, Everyday Modern Jewelry","edit",3.8,1.3,1.5,None,"yok / none","signet",2.15,"4558668281/slim-gold-signet-ring-dainty-minimalist"),
(84,4549046924,"Slim Rectangular Signet Ring | Minimalist Gold Band, Everyday Stackable Jewelry","edit",6,1,1.5,None,"yok / none","signet",2.72,"4549046924/slim-rectangular-signet-ring-minimalist"),
(85,4556239169,"Square Black Onyx Signet Ring | Minimalist Geometric Mens Jewelry","edit",8.7,2.5,1.5,None,"tekil/aksan tas","signet",4.62,"4556239169/black-onyx-square-signet-ring-mens"),
(86,4543742514,"Square Signet Ring | 10mm Minimalist Gold Band | Unisex Pinky Ring","edit",10,2.5,1.5,10,"yok / none","signet",5.07,"4543742514/square-signet-ring-10mm-minimalist-gold"),
(87,4550942827,"Stackable Gold Ring Set | Slim Charlotte Bands, Dainty Everyday Jewelry","edit",1.50,None,1.5,None,"yok / none","band",1.48,"4550942827/stackable-gold-ring-set-slim-charlotte"),
(88,4538023253,"Stevie Ring | Bold 6.5mm Wide Band | Modern Minimalist Statement Jewelry","edit",1.5,None,1.5,6.5,"yok / none","band",1.48,"4538023253/stevie-ring-65mm-statement-ring-gold"),
(89,4547283007,"Thin Gold Dome Ring | Minimalist Wedding Band, Everyday Stacking Jewelry","edit",6.2,2,1.5,None,"yok / none","dome",2.91,"4547283007/thin-gold-dome-ring-minimalist-wedding"),
(90,4543248600,"Twisted Duo Ring Set | Smooth and Twist Bands, Minimalist Wedding Jewelry","edit",1.5,None,1.5,None,"yok / none","band",1.48,"4543248600/twisted-duo-ring-set-smooth-and-twist"),
(91,4546206240,"Twisted Rope Band Ring | Minimalist Stacking Ring, Everyday Jewelry","edit",1.5,None,1.5,None,"yok / none","band",1.48,"4546206240/twisted-rope-band-ring-minimalist"),
(92,4550306442,"Ziggy Wave Ring | Sculptural Gold Band, Minimalist Stackable Jewelry","edit",4,1.3,1.5,None,"yok / none","band",2.22,"4550306442/ziggy-wave-ring-sculptural-gold-band"),
(93,4558671043,"10K/14k Solid Gold Wishbone Ring, 1mm Stacking Band with Clear Stone Accents","active",1,1,1.5,None,"tekil/aksan tas","band",0.99,"4558671043/dainty-gold-wishbone-stacking-ring-thin"),
]

CONFLICT = {4538023253,4543233648,4552128868,4552138588,4552078311,4553159638,4552114869,4543752254,4552097077}

wb = Workbook()

# ---------------- Sheet 1: cost request ----------------
ws = wb.active
ws.title = "Maliyet Talebi"
HEAD = [
 ("#",5),("Etsy ID",13),("Urun / Product",52),("Durum",9),("Form",9),
 ("Genislik mm\n(cikarim)",12),("Shank mm",10),("Kalinlik mm",11),
 ("Basliktaki mm",12),("KONTROL",11),("Tas (aciklamadan)",20),
 ("GRAM 14K\n(US 7)",12),("ISCILIK USD",12),("TAS+MIHLAMA USD",16),
 ("DOKUM+DIGER USD",16),("URETICI NOTU",30),
 ("Metal maliyeti USD",17),("TOPLAM MALIYET USD",18),
 ("Bizim gram tahminimiz",19),("Etsy link",46),
]
ws.append([h for h,_ in HEAD])
for i,(h,w) in enumerate(HEAD, start=1):
    c = ws.cell(row=1, column=i)
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    c.fill = FILLHDR if 12 <= i <= 16 else HDRFILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 34

for (n,eid,title,status,w,shank,thick,tmm,stone,form,estg,slug) in D:
    r = ws.max_row + 1
    conflict = eid in CONFLICT
    ws.cell(row=r,column=1,value=n)
    ws.cell(row=r,column=2,value=eid)
    ws.cell(row=r,column=3,value=title)
    ws.cell(row=r,column=4,value="Aktif" if status=="active" else "Pasif")
    ws.cell(row=r,column=5,value=form)
    ws.cell(row=r,column=6,value=w)
    ws.cell(row=r,column=7,value=shank)
    ws.cell(row=r,column=8,value=thick)
    ws.cell(row=r,column=9,value=tmm)
    ws.cell(row=r,column=10,value="GENISLIGI TEYIT ET" if conflict else "")
    ws.cell(row=r,column=11,value=stone)
    for col in (12,13,14,15,16):
        cc = ws.cell(row=r,column=col,value=None); cc.fill = YELLOW
    ws.cell(row=r,column=17,value=f"=IF($L{r}=\"\",\"\",$L{r}*Varsayimlar!$B$5*(1+Varsayimlar!$B$7))")
    ws.cell(row=r,column=18,value=f"=IF($L{r}=\"\",\"\",$Q{r}+N($M{r})+N($N{r})+N($O{r}))")
    ws.cell(row=r,column=19,value=estg).fill = REFFILL
    ws.cell(row=r,column=20,value=f"https://www.etsy.com/listing/{slug}")
    if conflict:
        for col in (6,9,10):
            ws.cell(row=r,column=col).fill = WARNFILL
    for col in range(1,21):
        cell = ws.cell(row=r,column=col)
        cell.border = BORDER
        if cell.font is None or cell.font.name != FONT:
            cell.font = Font(name=FONT, size=10)
        if col in (17,18):
            cell.number_format = '$#,##0.00'
        if col in (13,14,15):
            cell.number_format = '$#,##0.00'
        if col in (12,19):
            cell.number_format = '0.00'
        if col == 3:
            cell.alignment = Alignment(wrap_text=False, vertical="center")
        elif col in (10,11):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.freeze_panes = "D2"
ws.auto_filter.ref = f"A1:T{ws.max_row}"

note = ws.max_row + 2
ws.cell(row=note,column=3,value="SARI hucreler uretici tarafindan doldurulacak. KIRMIZI 'KONTROL' satirlarinda aciklama ile baslik farkli mm veriyor - genislik teyit edilmeli.").font = Font(name=FONT, bold=True, size=10, color="C00000")
ws.cell(row=note+1,column=3,value="Gram, 14K / US 7 referansinda istenir. Diger karat ve bedenler bu referanstan panelde olceklenir.").font = Font(name=FONT, italic=True, size=9)
ws.cell(row=note+2,column=3,value="Kaynak: olculer Etsy urun aciklamalarindan otomatik cikarildi (Ophir Gold USA, panel DB, 93 kayit, 2026-08-28).").font = Font(name=FONT, italic=True, size=9)

# ---------------- Sheet 2: assumptions ----------------
a = wb.create_sheet("Varsayimlar")
a.column_dimensions["A"].width = 38; a.column_dimensions["B"].width = 14; a.column_dimensions["C"].width = 68
a["A1"] = "Varsayimlar / Assumptions"; a["A1"].font = Font(name=FONT, bold=True, size=13)
rows = [
 ("Girdi","Deger","Kaynak / not"),
 ("10K altin alis (USD/gram)",65,"Panel > Ayarlar > Altin (gold_settings.purchase_price_10k_cents=6500)"),
 ("14K altin alis (USD/gram)",101,"Panel > Ayarlar > Altin (gold_settings.purchase_price_14k_cents=10100)"),
 ("18K altin alis (USD/gram)",130,"VARSAYIM - panelde 18K kaydi YOK. 14K'nin saflik oraniyla turetildi: 101/0.583*0.750."),
 ("Fire / kayip orani",0.07,"EON v4 motoru (lib/pricing/gold-index.ts, V4.fire). Uretici kendi firesini veriyorsa onu kullanin."),
 ("Referans beden","US 7","Gram talebi bu bedende. Panel diger bedenlere olcekler."),
 ("Referans kalinlik (mm)",1.5,"93 listingin 63'unde aciklamada yazili; gozlenen araligin tamami 1.50-1.60 mm."),
]
for i,(k,v,src) in enumerate(rows, start=3):
    a.cell(row=i,column=1,value=k).font = Font(name=FONT, bold=(i==3), size=10)
    c = a.cell(row=i,column=2,value=v)
    c.font = Font(name=FONT, bold=True, size=10) if i==3 else BLUE
    if i>3: c.fill = YELLOW
    if k.startswith("Fire"): c.number_format = '0.0%'
    elif isinstance(v,(int,float)): c.number_format = '$#,##0'
    a.cell(row=i,column=3,value=src).font = Font(name=FONT, italic=True, size=9)
    a.cell(row=i,column=3).alignment = Alignment(wrap_text=True, vertical="center")
a["A11"] = "Metal maliyeti = gram x 14K alis x (1 + fire).  Toplam maliyet = metal + iscilik + tas/mihlama + dokum/diger."
a["A11"].font = Font(name=FONT, italic=True, size=10)
a["A12"] = "SATIS fiyati bu sayfada YOK - once gercek maliyet toplanir, fiyat sonra panelde kurulur."
a["A12"].font = Font(name=FONT, bold=True, size=10, color="C00000")

# ---------------- Sheet 3: how to fill ----------------
h = wb.create_sheet("Nasil Doldurulur")
h.column_dimensions["A"].width = 26; h.column_dimensions["B"].width = 96
h["A1"] = "Nasil doldurulur / How to fill"; h["A1"].font = Font(name=FONT, bold=True, size=13)
guide = [
 ("Amac","93 listing icin GERCEK maliyet girdilerini ureticiden almak. Panelde su an gram YOK; tum listingler tek tip 366 USD fiyatta."),
 ("Doldurulacak alanlar","Yalnizca SARI hucreler: GRAM 14K (US 7), ISCILIK USD, TAS+MIHLAMA USD, DOKUM+DIGER USD, URETICI NOTU."),
 ("Gram","14 ayar ve US 7 beden icin parca basi net gram. Tek referans yeter; diger karat/bedenler panelde hesaplanir."),
 ("Iscilik","Parca basi iscilik (USD). Tas mihlama HARIC - o ayri sutunda."),
 ("Tas + mihlama","Tas bedeli + mihlama iscilligi toplami (USD). Tas yoksa 0 yazin."),
 ("Dokum + diger","Dokum, cila, kaplama vb. varsa (USD). Yoksa 0."),
 ("KONTROL sutunu","Kirmizi isaretli 9 satirda urun aciklamasi ile baslik FARKLI mm veriyor. Bu satirlarda dogru genisligi teyit edip nota yazin."),
 ("Hesaplanan sutunlar","'Metal maliyeti' ve 'TOPLAM MALIYET' otomatik hesaplanir - elle doldurmayin."),
 ("Bizim gram tahminimiz","Sadece BUYUKLUK FIKRI icin. Olcuden hesaplanan kaba tahmindir, dogrulanmis degildir - referans almayin."),
]
r = 3
for k,v in guide:
    h.cell(row=r,column=1,value=k).font = Font(name=FONT, bold=True, size=10)
    c = h.cell(row=r,column=2,value=v); c.font = Font(name=FONT, size=10); c.alignment = Alignment(wrap_text=True, vertical="center")
    h.row_dimensions[r].height = 30
    r += 1

r += 1
h.cell(row=r,column=1,value="ORNEK SATIR (beklenen format)").font = Font(name=FONT, bold=True, size=11, color="C00000")
r += 1
ex_h = ["Urun","GRAM 14K (US 7)","ISCILIK USD","TAS+MIHLAMA USD","DOKUM+DIGER USD","URETICI NOTU"]
for i,x in enumerate(ex_h, start=1):
    c = h.cell(row=r,column=i,value=x); c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10); c.fill = HDRFILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)
r += 1
ex = ["2mm Smooth Gold Band Ring", 2.05, 100, 0, 12, "Duz band, tassiz. US7 net gram."]
for i,x in enumerate(ex, start=1):
    c = h.cell(row=r,column=i,value=x); c.font = BLUE; c.fill = YELLOW; c.border = BORDER
    if i in (3,4,5): c.number_format = '$#,##0.00'
    if i == 2: c.number_format = '0.00'
for col in "CDEF":
    h.column_dimensions[col].width = 17

out = "/home/user/JadeGoldNYC/docs/ophir/ophir-maliyet-talebi.xlsx"
wb.save(out)
print("saved", out)
print("rows:", len(D), "conflicts:", len(CONFLICT))
print("checksum widths:", round(sum(d[4] for d in D),2), "| est_g14:", round(sum(d[10] for d in D),2))
